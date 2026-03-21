"""
Build gold standard annotation workbook.

One index sheet + one sheet per exercise.
Each exercise sheet shows all LLM output with editable correction columns.

Usage:
    python3 evals/build_gold_sheet.py
    python3 evals/build_gold_sheet.py --output evals/gold_annotation.xlsx
"""

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

_HERE = Path(__file__).parent
_ROOT = _HERE.parent
ENRICHED_DIR = _ROOT / "sources" / "free-exercise-db" / "enriched"

# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------
C_HEADER_BG = "1F3864"     # dark navy — sheet header band
C_HEADER_FG = "FFFFFF"
C_SECTION_BG = "D6E4F0"   # light blue — section labels
C_LLM_BG    = "EBF5FB"    # very light blue — LLM value cells (read-only feel)
C_EDIT_BG   = "FDFEFE"    # near-white — editable correction cells
C_TABLE_HDR = "2C3E50"    # dark for muscle table header
C_INDEX_HDR = "1F3864"

# Status fill colours
STATUS_FILLS = {
    "Pending":  PatternFill("solid", fgColor="F0F0F0"),
    "Accepted": PatternFill("solid", fgColor="D5F5E3"),
    "Modified": PatternFill("solid", fgColor="FEF9E7"),
    "Flagged":  PatternFill("solid", fgColor="FADBD8"),
}

STATUS_OPTIONS = '"Pending,Accepted,Modified,Flagged"'
DEGREE_OPTIONS = '"PrimeMover,Synergist,Stabilizer,PassiveTarget"'
ROW_STATUS_OPTIONS = '"Pending,Accept,Reject,Modify"'

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _dv(formula: str, sqref: str) -> DataValidation:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=False)
    dv.sqref = sqref
    return dv


def _safe_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no special chars."""
    bad = r"\/*?:[]\'"
    for ch in bad:
        name = name.replace(ch, "_")
    return name[:31]


def _fmt_list(val) -> str:
    if not val:
        return ""
    if isinstance(val, list):
        return ", ".join(str(v) for v in val)
    return str(val)


def _fmt_bool(val) -> str:
    if val is None:
        return ""
    return "TRUE" if val else "FALSE"


# ---------------------------------------------------------------------------
# Exercise sheet builder
# ---------------------------------------------------------------------------
def build_exercise_sheet(ws, ex: dict) -> None:
    ws.freeze_panes = "A8"  # freeze header block

    # Column widths
    ws.column_dimensions["A"].width = 22   # field label
    ws.column_dimensions["B"].width = 40   # LLM value
    ws.column_dimensions["C"].width = 40   # corrected value
    ws.column_dimensions["D"].width = 14   # status / degree
    ws.column_dimensions["E"].width = 20   # corrected muscle (involvements table)
    ws.column_dimensions["F"].width = 18   # corrected degree
    ws.column_dimensions["G"].width = 14   # row status

    row = 1

    # ---- Title band --------------------------------------------------------
    ws.merge_cells(f"A{row}:G{row}")
    cell = ws[f"A{row}"]
    cell.value = ex.get("name", ex["id"])
    cell.font = _font(bold=True, color=C_HEADER_FG, size=13)
    cell.fill = _fill(C_HEADER_BG)
    cell.alignment = _align("center")
    ws.row_dimensions[row].height = 22
    row += 1

    ws.merge_cells(f"A{row}:G{row}")
    cell = ws[f"A{row}"]
    cell.value = f"id: {ex['id']}   |   category: {ex.get('category', '')}   |   level: {ex.get('level', '')}"
    cell.font = _font(color=C_HEADER_FG, size=10)
    cell.fill = _fill(C_HEADER_BG)
    cell.alignment = _align("center")
    row += 1

    row += 1  # spacer

    # ---- Column header for field block ------------------------------------
    for col, label in enumerate(["Field", "LLM Output", "Corrected Value", "Status"], start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = _font(bold=True, color=C_HEADER_FG)
        cell.fill = _fill(C_TABLE_HDR)
        cell.alignment = _align("center")
        cell.border = BORDER
    row += 1

    # ---- Scalar / list fields ---------------------------------------------
    scalar_fields = [
        ("Movement Patterns",      _fmt_list(ex.get("movement_patterns"))),
        ("Primary Joint Actions",  _fmt_list(ex.get("primary_joint_actions"))),
        ("Supporting Joint Actions", _fmt_list(ex.get("supporting_joint_actions"))),
        ("Is Compound",            _fmt_bool(ex.get("is_compound"))),
        ("Is Unilateral",          _fmt_bool(ex.get("is_unilateral"))),
        ("Training Modalities",    _fmt_list(ex.get("training_modalities"))),
    ]

    field_status_rows = []
    for label, llm_val in scalar_fields:
        ws.cell(row=row, column=1, value=label).font = _font(bold=True)
        ws.cell(row=row, column=1).fill = _fill(C_SECTION_BG)
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=1).alignment = _align()

        llm_cell = ws.cell(row=row, column=2, value=llm_val)
        llm_cell.fill = _fill(C_LLM_BG)
        llm_cell.border = BORDER
        llm_cell.alignment = _align(wrap=True)

        corr_cell = ws.cell(row=row, column=3, value="")
        corr_cell.fill = _fill(C_EDIT_BG)
        corr_cell.border = BORDER
        corr_cell.alignment = _align(wrap=True)

        status_cell = ws.cell(row=row, column=4, value="Pending")
        status_cell.fill = STATUS_FILLS["Pending"]
        status_cell.border = BORDER
        status_cell.alignment = _align("center")

        field_status_rows.append(f"D{row}")
        row += 1

    # Add status dropdown for scalar field rows
    for sqref in field_status_rows:
        ws.add_data_validation(_dv(STATUS_OPTIONS, sqref))

    row += 1  # spacer

    # ---- Muscle involvements table ----------------------------------------
    ws.merge_cells(f"A{row}:G{row}")
    section_cell = ws[f"A{row}"]
    section_cell.value = "Muscle Involvements"
    section_cell.font = _font(bold=True, color=C_HEADER_FG)
    section_cell.fill = _fill(C_TABLE_HDR)
    section_cell.alignment = _align("center")
    row += 1

    # Table header
    inv_headers = ["Muscle", "Degree", "Corrected Muscle", "Corrected Degree", "", "Row Status"]
    for col, label in enumerate(inv_headers, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = _font(bold=True, color=C_HEADER_FG)
        cell.fill = _fill(C_HEADER_BG)
        cell.alignment = _align("center")
        cell.border = BORDER
    row += 1

    inv_start = row
    for inv in ex.get("muscle_involvements", []):
        muscle = inv.get("muscle", "")
        degree = inv.get("degree", "")

        ws.cell(row=row, column=1, value=muscle).fill = _fill(C_LLM_BG)
        ws.cell(row=row, column=1).border = BORDER

        ws.cell(row=row, column=2, value=degree).fill = _fill(C_LLM_BG)
        ws.cell(row=row, column=2).border = BORDER
        ws.cell(row=row, column=2).alignment = _align("center")

        ws.cell(row=row, column=3, value="").fill = _fill(C_EDIT_BG)
        ws.cell(row=row, column=3).border = BORDER

        ws.cell(row=row, column=4, value="").fill = _fill(C_EDIT_BG)
        ws.cell(row=row, column=4).border = BORDER
        ws.cell(row=row, column=4).alignment = _align("center")

        status_cell = ws.cell(row=row, column=6, value="Pending")
        status_cell.fill = STATUS_FILLS["Pending"]
        status_cell.border = BORDER
        status_cell.alignment = _align("center")

        row += 1

    inv_end = row - 1

    # Dropdowns for involvement table
    if inv_end >= inv_start:
        ws.add_data_validation(_dv(DEGREE_OPTIONS, f"D{inv_start}:D{inv_end}"))
        ws.add_data_validation(_dv(ROW_STATUS_OPTIONS, f"F{inv_start}:F{inv_end}"))

    row += 2  # spacer

    # ---- Overall status + notes -------------------------------------------
    ws.merge_cells(f"A{row}:G{row}")
    overall_label = ws[f"A{row}"]
    overall_label.value = "Overall Exercise Status"
    overall_label.font = _font(bold=True, color=C_HEADER_FG)
    overall_label.fill = _fill(C_TABLE_HDR)
    overall_label.alignment = _align("center")
    row += 1

    ws.cell(row=row, column=1, value="Status").font = _font(bold=True)
    ws.cell(row=row, column=1).fill = _fill(C_SECTION_BG)
    ws.cell(row=row, column=1).border = BORDER

    overall_status = ws.cell(row=row, column=2, value="Pending")
    overall_status.fill = STATUS_FILLS["Pending"]
    overall_status.border = BORDER
    overall_status.alignment = _align("center")
    ws.add_data_validation(_dv(STATUS_OPTIONS, f"B{row}"))
    row += 1

    ws.cell(row=row, column=1, value="Notes").font = _font(bold=True)
    ws.cell(row=row, column=1).fill = _fill(C_SECTION_BG)
    ws.cell(row=row, column=1).border = BORDER

    notes_cell = ws.cell(row=row, column=2, value="")
    notes_cell.fill = _fill(C_EDIT_BG)
    notes_cell.border = BORDER
    notes_cell.alignment = _align(wrap=True)
    ws.row_dimensions[row].height = 40
    ws.merge_cells(f"B{row}:G{row}")


# ---------------------------------------------------------------------------
# Index sheet builder
# ---------------------------------------------------------------------------
def build_index_sheet(ws, exercises: list) -> None:
    ws.freeze_panes = "A3"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 40

    row = 1

    # Title
    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = "Gold Standard Annotation — Index"
    title.font = _font(bold=True, color=C_HEADER_FG, size=13)
    title.fill = _fill(C_INDEX_HDR)
    title.alignment = _align("center")
    ws.row_dimensions[1].height = 22
    row += 1

    # Column headers
    headers = ["#", "Exercise", "Category", "Movement Patterns", "Primary Joint Actions",
               "Is Compound", "Is Unilateral", "Status", "Notes"]
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = _font(bold=True, color=C_HEADER_FG)
        cell.fill = _fill(C_TABLE_HDR)
        cell.alignment = _align("center")
        cell.border = BORDER
    row += 1

    for i, ex in enumerate(exercises, start=1):
        ws.cell(row=row, column=1, value=i).alignment = _align("center")
        ws.cell(row=row, column=1).border = BORDER

        name_cell = ws.cell(row=row, column=2, value=ex.get("name", ex["id"]))
        name_cell.border = BORDER
        # Hyperlink to exercise sheet
        sheet_name = _safe_sheet_name(ex["id"])
        name_cell.hyperlink = f"#{sheet_name}!A1"
        name_cell.font = Font(color="1F497D", underline="single", name="Calibri", size=11)

        ws.cell(row=row, column=3, value=ex.get("category", "")).border = BORDER
        ws.cell(row=row, column=3).alignment = _align("center")

        ws.cell(row=row, column=4, value=_fmt_list(ex.get("movement_patterns"))).border = BORDER
        ws.cell(row=row, column=4).alignment = _align(wrap=True)

        ws.cell(row=row, column=5, value=_fmt_list(ex.get("primary_joint_actions"))).border = BORDER
        ws.cell(row=row, column=5).alignment = _align(wrap=True)

        ws.cell(row=row, column=6, value=_fmt_bool(ex.get("is_compound"))).border = BORDER
        ws.cell(row=row, column=6).alignment = _align("center")

        ws.cell(row=row, column=7, value=_fmt_bool(ex.get("is_unilateral"))).border = BORDER
        ws.cell(row=row, column=7).alignment = _align("center")

        status_cell = ws.cell(row=row, column=8, value="Pending")
        status_cell.fill = STATUS_FILLS["Pending"]
        status_cell.border = BORDER
        status_cell.alignment = _align("center")
        ws.add_data_validation(_dv(STATUS_OPTIONS, f"H{row}"))

        ws.cell(row=row, column=9, value="").border = BORDER
        ws.cell(row=row, column=9).alignment = _align(wrap=True)

        ws.row_dimensions[row].height = 30
        row += 1


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Build gold standard annotation xlsx")
    parser.add_argument("--output", default=str(_HERE / "gold_annotation.xlsx"))
    args = parser.parse_args()

    enriched_files = sorted(ENRICHED_DIR.glob("*.json"))
    if not enriched_files:
        print("No enriched files found. Run enrich.py first.")
        return
    exercises = [json.loads(p.read_text()) for p in enriched_files]
    exercises = sorted(exercises, key=lambda e: e.get("name", e["id"]).lower())

    wb = Workbook()

    # Index sheet
    ws_index = wb.active
    ws_index.title = "Index"
    build_index_sheet(ws_index, exercises)

    # One sheet per exercise
    for ex in exercises:
        sheet_name = _safe_sheet_name(ex["id"])
        ws = wb.create_sheet(title=sheet_name)
        build_exercise_sheet(ws, ex)

    out = Path(args.output)
    wb.save(out)
    print(f"Wrote {len(exercises)} exercises → {out}")


if __name__ == "__main__":
    main()
