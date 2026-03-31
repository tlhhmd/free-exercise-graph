"""
Evaluation scorer for enrichment pipeline output.

Scores predicted exercise attributes against a human-annotated gold standard.
See ADR-109 for methodology decisions.

Usage:
    python3 evals/eval.py --gold evals/gold_annotation.xlsx
    python3 evals/eval.py --gold evals/gold_annotation.xlsx --verbose
    python3 evals/eval.py --gold evals/gold_annotation.xlsx --field movement_pattern

Gold standard format: the xlsx produced by build_gold_sheet.py, with correction
columns filled in by a human annotator. Rows with status "Pending" are skipped.

Scoring methodology (ADR-109):
  Muscle involvements — three dimensions:
    strict_f1:    (muscle, degree) pair must match exactly
    muscle_f1:    muscle name match, degree ignored
    degree_acc:   for correctly identified muscles, fraction with correct degree

  All other multi-value fields (movement_pattern, primary_joint_action, etc.):
    exact set match F1

  Boolean fields (is_compound, is_unilateral):
    binary accuracy
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

_ROOT = Path(__file__).parent.parent
_DB_PATH = _ROOT / "pipeline" / "pipeline.db"

# ─── Metric helpers ──────────────────────────────────────────────────────────

def _prf(tp: int, fp: int, fn: int) -> dict[str, float]:
    p = tp / (tp + fp) if (tp + fp) else 0.0
    r = tp / (tp + fn) if (tp + fn) else 0.0
    f = 2 * p * r / (p + r) if (p + r) else 0.0
    return {"precision": p, "recall": r, "f1": f, "tp": tp, "fp": fp, "fn": fn}


def set_f1(pred: set, gold: set) -> dict[str, float]:
    tp = len(pred & gold)
    fp = len(pred - gold)
    fn = len(gold - pred)
    return _prf(tp, fp, fn)


def muscle_scores(
    pred_involvements: list[dict],
    gold_involvements: list[dict],
) -> dict[str, Any]:
    """
    Three-dimension muscle scoring (ADR-109):
      strict_f1   — (muscle, degree) pair exact match
      muscle_f1   — muscle name match, degree ignored
      degree_acc  — degree correct, conditional on correct muscle identification
    """
    pred_pairs = {(i["muscle"], i["degree"]) for i in pred_involvements}
    gold_pairs = {(i["muscle"], i["degree"]) for i in gold_involvements}
    pred_names = {i["muscle"] for i in pred_involvements}
    gold_names = {i["muscle"] for i in gold_involvements}

    strict = set_f1(pred_pairs, gold_pairs)
    muscle = set_f1(pred_names, gold_names)

    # Degree accuracy: among muscles present in both pred and gold,
    # what fraction have the correct degree?
    shared_muscles = pred_names & gold_names
    degree_correct = 0
    degree_total = 0
    if shared_muscles:
        gold_degree_map = {i["muscle"]: i["degree"] for i in gold_involvements}
        pred_degree_map = {i["muscle"]: i["degree"] for i in pred_involvements}
        for m in shared_muscles:
            degree_total += 1
            if pred_degree_map.get(m) == gold_degree_map.get(m):
                degree_correct += 1

    degree_acc = degree_correct / degree_total if degree_total else None

    return {
        "strict_f1": strict["f1"],
        "strict_precision": strict["precision"],
        "strict_recall": strict["recall"],
        "muscle_f1": muscle["f1"],
        "muscle_precision": muscle["precision"],
        "muscle_recall": muscle["recall"],
        "degree_acc": degree_acc,
        "degree_correct": degree_correct,
        "degree_total": degree_total,
    }


# ─── Gold standard loader ─────────────────────────────────────────────────────

def _parse_csv_cell(value: str | None) -> list[str]:
    """Parse a comma-separated cell value into a list of stripped strings."""
    if not value:
        return []
    return [v.strip() for v in str(value).split(",") if v.strip()]


def _bool_cell(value: str | None) -> bool | None:
    if value is None:
        return None
    v = str(value).strip().upper()
    if v in ("TRUE", "YES", "1"):
        return True
    if v in ("FALSE", "NO", "0"):
        return False
    return None


def load_gold(xlsx_path: Path) -> dict[str, dict]:
    """
    Load annotated gold standard from xlsx.

    Returns {exercise_id: gold_record} for all exercises whose
    overall status is not "Pending".

    Gold record shape:
      {
        "id": str,
        "movement_patterns": [str, ...],
        "primary_joint_actions": [str, ...],
        "supporting_joint_actions": [str, ...],
        "is_compound": bool | None,
        "is_unilateral": bool | None,
        "training_modalities": [str, ...],
        "muscle_involvements": [{"muscle": str, "degree": str}, ...],
      }
    """
    wb = load_workbook(xlsx_path, data_only=True)
    gold: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        if sheet_name == "Index":
            continue
        ws = wb[sheet_name]

        # Exercise id is in row 2, col A value like "id: ffdb_... | ..."
        id_row = ws.cell(row=2, column=1).value or ""
        exercise_id = None
        for part in str(id_row).split("|"):
            if "id:" in part:
                exercise_id = part.split("id:")[-1].strip()
                break
        if not exercise_id:
            continue

        # Overall status is written by build_gold_sheet.py in a known row.
        # We scan for the "Overall Exercise Status" section marker and read
        # the status cell two rows below it.
        overall_status = "Pending"
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "Overall Exercise Status":
                    status_cell = ws.cell(row=cell.row + 2, column=2)
                    overall_status = str(status_cell.value or "Pending").strip()
                    break

        if overall_status == "Pending":
            continue

        # Read scalar fields. They appear in a fixed order after the column
        # header row. We scan for them by label in column A.
        scalar: dict[str, Any] = {}
        field_map = {
            "Movement Patterns": "movement_patterns",
            "Primary Joint Actions": "primary_joint_actions",
            "Supporting Joint Actions": "supporting_joint_actions",
            "Is Compound": "is_compound",
            "Is Unilateral": "is_unilateral",
            "Training Modalities": "training_modalities",
        }

        for row in ws.iter_rows():
            label_cell = row[0]
            label = str(label_cell.value or "").strip()
            if label not in field_map:
                continue
            # Column C is the corrected value; fall back to column B (LLM value)
            # if correction is blank.
            corrected = ws.cell(row=label_cell.row, column=3).value
            llm_val = ws.cell(row=label_cell.row, column=2).value
            value = corrected if corrected not in (None, "") else llm_val
            field_key = field_map[label]
            if field_key in ("is_compound", "is_unilateral"):
                scalar[field_key] = _bool_cell(value)
            else:
                scalar[field_key] = _parse_csv_cell(value)

        # Read muscle involvement table. Scan for "Muscle Involvements" header,
        # then read rows until we hit an empty muscle cell.
        involvements: list[dict] = []
        in_inv_table = False
        past_header = False

        for row in ws.iter_rows():
            label = str(row[0].value or "").strip()
            if label == "Muscle Involvements":
                in_inv_table = True
                continue
            if not in_inv_table:
                continue
            # Skip the column header row ("Muscle", "Degree", ...)
            if not past_header:
                past_header = True
                continue
            # Corrected muscle in col C (index 2), corrected degree in col D (index 3).
            # Fall back to LLM values in col A (index 0) / col B (index 1).
            corr_muscle = row[2].value if len(row) > 2 else None
            corr_degree = row[3].value if len(row) > 3 else None
            llm_muscle = row[0].value
            llm_degree = row[1].value

            muscle = str(corr_muscle or llm_muscle or "").strip()
            degree = str(corr_degree or llm_degree or "").strip()

            if not muscle:
                break  # end of table
            involvements.append({"muscle": muscle, "degree": degree})

        gold[exercise_id] = {
            "id": exercise_id,
            **scalar,
            "muscle_involvements": involvements,
        }

    return gold


# ─── Prediction loader ────────────────────────────────────────────────────────

def load_predictions(exercise_ids: list[str], db_path: Path = _DB_PATH) -> dict[str, dict]:
    """
    Load enrichment predictions from pipeline.db for the given exercise IDs.

    Reads inferred_claims and resolved_claims (resolved takes precedence, matching
    the same precedence logic used by build.py).
    """
    import sqlite3

    conn = sqlite3.connect(db_path)
    preds: dict[str, dict] = {}

    for eid in exercise_ids:
        # Collect all claims for this entity; resolved overrides inferred
        claims: dict[str, list] = {}

        for predicate, value, qualifier in conn.execute(
            "SELECT predicate, value, qualifier FROM inferred_claims WHERE entity_id = ?",
            (eid,),
        ):
            claims.setdefault(predicate, []).append((value, qualifier))

        # Collect resolved claims per predicate, then override inferred per predicate
        resolved: dict[str, list] = {}
        for predicate, value, qualifier in conn.execute(
            "SELECT predicate, value, qualifier FROM resolved_claims WHERE entity_id = ?",
            (eid,),
        ):
            resolved.setdefault(predicate, []).append((value, qualifier))
        # Resolved overrides inferred at the predicate level (same logic as build.py)
        for predicate, values in resolved.items():
            claims[predicate] = values

        # Reshape into the eval format
        muscle_involvements = [
            {"muscle": v, "degree": q or ""}
            for v, q in claims.get("muscle", [])
        ]
        movement_patterns = [v for v, _ in claims.get("movement_pattern", [])]
        primary_jas = [v for v, _ in claims.get("primary_joint_action", [])]
        supporting_jas = [v for v, _ in claims.get("supporting_joint_action", [])]
        training_modalities = [v for v, _ in claims.get("training_modality", [])]

        is_compound_vals = [v for v, _ in claims.get("is_compound", [])]
        is_compound = _bool_cell(is_compound_vals[0]) if is_compound_vals else None

        laterality_vals = [v for v, _ in claims.get("laterality", [])]
        is_unilateral = (
            laterality_vals[0].lower() in ("unilateral", "ipsilateral", "contralateral")
            if laterality_vals else None
        )

        preds[eid] = {
            "id": eid,
            "movement_patterns": movement_patterns,
            "primary_joint_actions": primary_jas,
            "supporting_joint_actions": supporting_jas,
            "training_modalities": training_modalities,
            "is_compound": is_compound,
            "is_unilateral": is_unilateral,
            "muscle_involvements": muscle_involvements,
        }

    conn.close()
    return preds


# ─── Scoring ──────────────────────────────────────────────────────────────────

def score_exercise(pred: dict, gold: dict) -> dict[str, Any]:
    """Score a single exercise. Returns per-field metrics."""
    results: dict[str, Any] = {"id": gold["id"]}

    # Movement patterns
    pred_mp = set(pred.get("movement_patterns") or [])
    gold_mp = set(gold.get("movement_patterns") or [])
    results["movement_pattern"] = set_f1(pred_mp, gold_mp)

    # Primary joint actions
    pred_pja = set(pred.get("primary_joint_actions") or [])
    gold_pja = set(gold.get("primary_joint_actions") or [])
    results["primary_joint_action"] = set_f1(pred_pja, gold_pja)

    # Supporting joint actions
    pred_sja = set(pred.get("supporting_joint_actions") or [])
    gold_sja = set(gold.get("supporting_joint_actions") or [])
    results["supporting_joint_action"] = set_f1(pred_sja, gold_sja)

    # Training modalities
    pred_tm = set(pred.get("training_modalities") or [])
    gold_tm = set(gold.get("training_modalities") or [])
    results["training_modality"] = set_f1(pred_tm, gold_tm)

    # Boolean fields
    pred_compound = pred.get("is_compound")
    gold_compound = gold.get("is_compound")
    results["is_compound"] = (
        int(pred_compound == gold_compound)
        if gold_compound is not None else None
    )

    pred_uni = pred.get("is_unilateral")
    gold_uni = gold.get("is_unilateral")
    results["is_unilateral"] = (
        int(pred_uni == gold_uni)
        if gold_uni is not None else None
    )

    # Muscle involvements — three dimensions
    pred_inv = pred.get("muscle_involvements") or []
    gold_inv = gold.get("muscle_involvements") or []
    results["muscle"] = muscle_scores(pred_inv, gold_inv)

    return results


def aggregate(per_exercise: list[dict]) -> dict[str, Any]:
    """Macro-average metrics across all scored exercises."""

    def _macro_f1(field: str) -> dict[str, float]:
        ps, rs, fs = [], [], []
        for ex in per_exercise:
            m = ex.get(field)
            if m and isinstance(m, dict):
                ps.append(m["precision"])
                rs.append(m["recall"])
                fs.append(m["f1"])
        if not fs:
            return {"precision": 0.0, "recall": 0.0, "f1": 0.0, "n": 0}
        return {
            "precision": sum(ps) / len(ps),
            "recall": sum(rs) / len(rs),
            "f1": sum(fs) / len(fs),
            "n": len(fs),
        }

    def _bool_acc(field: str) -> dict[str, float]:
        vals = [ex[field] for ex in per_exercise if ex.get(field) is not None]
        if not vals:
            return {"accuracy": 0.0, "n": 0}
        return {"accuracy": sum(vals) / len(vals), "n": len(vals)}

    agg: dict[str, Any] = {
        "n": len(per_exercise),
        "movement_pattern": _macro_f1("movement_pattern"),
        "primary_joint_action": _macro_f1("primary_joint_action"),
        "supporting_joint_action": _macro_f1("supporting_joint_action"),
        "training_modality": _macro_f1("training_modality"),
        "is_compound": _bool_acc("is_compound"),
        "is_unilateral": _bool_acc("is_unilateral"),
    }

    # Muscle: aggregate all three dimensions
    strict_ps, strict_rs, strict_fs = [], [], []
    muscle_ps, muscle_rs, muscle_fs = [], [], []
    degree_accs: list[float] = []

    for ex in per_exercise:
        m = ex.get("muscle")
        if not m:
            continue
        strict_ps.append(m["strict_precision"])
        strict_rs.append(m["strict_recall"])
        strict_fs.append(m["strict_f1"])
        muscle_ps.append(m["muscle_precision"])
        muscle_rs.append(m["muscle_recall"])
        muscle_fs.append(m["muscle_f1"])
        if m["degree_acc"] is not None:
            degree_accs.append(m["degree_acc"])

    def _avg(lst: list[float]) -> float:
        return sum(lst) / len(lst) if lst else 0.0

    agg["muscle"] = {
        "strict_f1": _avg(strict_fs),
        "strict_precision": _avg(strict_ps),
        "strict_recall": _avg(strict_rs),
        "muscle_f1": _avg(muscle_fs),
        "muscle_precision": _avg(muscle_ps),
        "muscle_recall": _avg(muscle_rs),
        "degree_acc": _avg(degree_accs) if degree_accs else None,
        "n": len(strict_fs),
    }

    return agg


# ─── Reporting ────────────────────────────────────────────────────────────────

def _pct(v: float | None) -> str:
    if v is None:
        return "  —  "
    return f"{v * 100:.1f}%"


def print_report(agg: dict, per_exercise: list[dict], verbose: bool = False) -> None:
    n = agg["n"]
    print(f"\nEval Report — {n} exercise(s) scored")
    print("─" * 62)
    print(f"{'Field':<30} {'P':>7} {'R':>7} {'F1':>7}")
    print("─" * 62)

    set_fields = [
        ("movement_pattern",       "Movement Pattern"),
        ("primary_joint_action",   "Primary Joint Action"),
        ("supporting_joint_action","Supporting Joint Action"),
        ("training_modality",      "Training Modality"),
    ]
    for key, label in set_fields:
        m = agg[key]
        print(f"  {label:<28} {_pct(m['precision'])} {_pct(m['recall'])} {_pct(m['f1'])}")

    print()
    m = agg["muscle"]
    print(f"  {'Muscle (strict, w/ degree)':<28} {_pct(m['strict_precision'])} {_pct(m['strict_recall'])} {_pct(m['strict_f1'])}")
    print(f"  {'Muscle (name only)':<28} {_pct(m['muscle_precision'])} {_pct(m['muscle_recall'])} {_pct(m['muscle_f1'])}")
    print(f"  {'Degree accuracy (cond.)':<28}                   {_pct(m['degree_acc'])}")

    print()
    ca = agg["is_compound"]
    ua = agg["is_unilateral"]
    print(f"  {'Is Compound':<28}                   {_pct(ca['accuracy'])}  (n={ca['n']})")
    print(f"  {'Is Unilateral':<28}                   {_pct(ua['accuracy'])}  (n={ua['n']})")
    print("─" * 62)

    if verbose:
        print("\nPer-exercise breakdown:")
        for ex in per_exercise:
            eid = ex["id"]
            mp = ex["movement_pattern"]["f1"]
            ms = ex["muscle"]["strict_f1"]
            mn = ex["muscle"]["muscle_f1"]
            da = ex["muscle"]["degree_acc"]
            print(f"  {eid}")
            print(f"    movement_pattern f1={_pct(mp)}  muscle_strict={_pct(ms)}  muscle_name={_pct(mn)}  degree_acc={_pct(da)}")


# ─── CLI ─────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Score enrichment output against gold standard")
    parser.add_argument("--gold", required=True, help="Path to annotated gold_annotation.xlsx")
    parser.add_argument("--db", default=str(_DB_PATH), help="Path to pipeline.db")
    parser.add_argument("--verbose", "-v", action="store_true", help="Print per-exercise breakdown")
    parser.add_argument("--field", help="Score only this field (e.g. movement_pattern)")
    args = parser.parse_args()

    gold_path = Path(args.gold)
    if not gold_path.exists():
        sys.exit(f"Gold file not found: {gold_path}")

    db_path = Path(args.db)
    if not db_path.exists():
        sys.exit(f"pipeline.db not found: {db_path}")

    print(f"Loading gold standard from {gold_path}...")
    gold = load_gold(gold_path)
    if not gold:
        sys.exit("No annotated exercises found (all Pending, or gold file is empty).")
    print(f"  {len(gold)} annotated exercise(s) loaded.")

    print(f"Loading predictions from {db_path}...")
    preds = load_predictions(list(gold.keys()), db_path)
    missing = [eid for eid in gold if eid not in preds]
    if missing:
        print(f"  Warning: {len(missing)} exercise(s) in gold have no prediction file: {missing[:5]}")

    per_exercise = []
    for eid, gold_rec in gold.items():
        pred = preds.get(eid, {})
        per_exercise.append(score_exercise(pred, gold_rec))

    agg = aggregate(per_exercise)
    print_report(agg, per_exercise, verbose=args.verbose)


if __name__ == "__main__":
    main()
