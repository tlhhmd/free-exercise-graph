#!/usr/bin/env python3
"""
fetch.py — Download the Functional Fitness Exercise Database and convert to CSV.

Downloads the Excel file from strengthtoovercome.com, saves it to raw/exercises.xlsx,
and exports each sheet as raw/{SheetName}.csv.

Usage:
    python3 sources/functional-fitness-db/fetch.py

Source URL is documented in sources/functional-fitness-db/catalog.ttl
(dcat:downloadURL on the feg:SourceDistribution resource).

Credit: Strength to Overcome — https://strengthtoovercome.com/functional-fitness-exercise-database
"""

import csv
from pathlib import Path

import httpx
import openpyxl

SOURCE_URL = "https://strengthtoovercome.com/s/Functional-Fitness-Exercise-Database-version-29.xlsx"
RAW_DIR = Path(__file__).resolve().parent / "raw"
XLSX_DEST = RAW_DIR / "exercises.xlsx"


def main():
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Fetching {SOURCE_URL}")
    with httpx.Client(follow_redirects=True, timeout=60) as client:
        response = client.get(SOURCE_URL)
        response.raise_for_status()
    XLSX_DEST.write_bytes(response.content)
    print(f"Written: {XLSX_DEST} ({len(response.content):,} bytes)")

    print("\nConverting sheets to CSV...")
    wb = openpyxl.load_workbook(XLSX_DEST, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        csv_path = RAW_DIR / f"{sheet_name}.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
        print(f"  {sheet_name}.csv ({csv_path.stat().st_size:,} bytes)")
    wb.close()


if __name__ == "__main__":
    main()
